import os
import tempfile
import unittest
from unittest.mock import patch, MagicMock

from openpyxl import Workbook, load_workbook

import excel
from scrape import resolve_input_path


def _create_excel(path, headers, rows):
    """Helper: write a small Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


class TestReadExcelInput(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_basic_read(self):
        path = os.path.join(self.tmpdir, "input.xlsx")
        _create_excel(path, ["CodArticol", "Desc"], [["A100", "Widget"], ["B200", "Gadget"]])

        terms, rows, cols = excel.read_excel_input(path, "CodArticol")

        self.assertEqual(terms, ["A100", "B200"])
        self.assertEqual(len(rows), 2)
        self.assertEqual(cols, ["CodArticol", "Desc"])
        self.assertEqual(rows[0]["CodArticol"], "A100")
        self.assertEqual(rows[1]["Desc"], "Gadget")

    def test_skips_empty_id_rows(self):
        path = os.path.join(self.tmpdir, "input.xlsx")
        _create_excel(path, ["CodArticol", "Desc"], [["A100", "Widget"], [None, "Empty"], ["", "Blank"]])

        terms, rows, _ = excel.read_excel_input(path, "CodArticol")

        self.assertEqual(terms, ["A100"])
        self.assertEqual(len(rows), 1)

    def test_missing_id_column_raises(self):
        path = os.path.join(self.tmpdir, "input.xlsx")
        _create_excel(path, ["Name", "Desc"], [["A100", "Widget"]])

        with self.assertRaises(ValueError) as ctx:
            excel.read_excel_input(path, "CodArticol")
        self.assertIn("CodArticol", str(ctx.exception))

    def test_file_not_found_raises(self):
        with self.assertRaises(FileNotFoundError):
            excel.read_excel_input("/nonexistent/input.xlsx", "CodArticol")

    def test_empty_file_raises(self):
        path = os.path.join(self.tmpdir, "empty.xlsx")
        wb = Workbook()
        ws = wb.active
        # No rows at all — remove the default empty row by ensuring ws has nothing
        wb.save(path)

        with self.assertRaises(ValueError) as ctx:
            excel.read_excel_input(path, "CodArticol")
        self.assertIn("empty", str(ctx.exception).lower())


class TestExportScraperExcel(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_merge_two_matches(self):
        original_rows = [
            {"CodArticol": "A100", "Desc": "Widget"},
            {"CodArticol": "B200", "Desc": "Gadget"},
        ]
        original_columns = ["CodArticol", "Desc"]
        scraper_results = [
            {"CodArticol": "A100", "PretNormal": 10.5, "PretStocAD": 12.0},
            {"CodArticol": "B200", "PretNormal": 20.0, "PretStocAD": 22.0},
        ]
        out_path = os.path.join(self.tmpdir, "out.xlsx")

        result = excel.export_scraper_excel(original_rows, original_columns, scraper_results, "CodArticol", out_path)

        self.assertEqual(result, out_path)
        wb = load_workbook(out_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, ["CodArticol", "Desc", "PretNormal", "PretStocAD"])

        row1 = [c.value for c in ws[2]]
        self.assertEqual(row1, ["A100", "Widget", 10.5, 12.0])

        row2 = [c.value for c in ws[3]]
        self.assertEqual(row2, ["B200", "Gadget", 20.0, 22.0])
        wb.close()

    def test_partial_matches(self):
        original_rows = [
            {"CodArticol": "A100", "Desc": "Widget"},
            {"CodArticol": "B200", "Desc": "Gadget"},
        ]
        original_columns = ["CodArticol", "Desc"]
        scraper_results = [
            {"CodArticol": "A100", "PretNormal": 10.5, "PretStocAD": 12.0},
        ]
        out_path = os.path.join(self.tmpdir, "out.xlsx")

        excel.export_scraper_excel(original_rows, original_columns, scraper_results, "CodArticol", out_path)

        wb = load_workbook(out_path)
        ws = wb.active
        # Row with no scraper match should have empty price cells
        row2 = [c.value for c in ws[3]]
        self.assertEqual(row2[0], "B200")
        self.assertEqual(row2[1], "Gadget")
        self.assertIn(row2[2], ("", None))  # PretNormal empty
        self.assertIn(row2[3], ("", None))  # PretStocAD empty
        wb.close()

    def test_empty_input_returns_none(self):
        result = excel.export_scraper_excel([], ["CodArticol"], [], "CodArticol", "/tmp/out.xlsx")
        self.assertIsNone(result)

    def test_preserves_original_column_order(self):
        original_rows = [{"Z_Col": "z", "A_Col": "a", "CodArticol": "X1"}]
        original_columns = ["Z_Col", "A_Col", "CodArticol"]
        scraper_results = [{"CodArticol": "X1", "Price": 5.0}]
        out_path = os.path.join(self.tmpdir, "out.xlsx")

        excel.export_scraper_excel(original_rows, original_columns, scraper_results, "CodArticol", out_path)

        wb = load_workbook(out_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, ["Z_Col", "A_Col", "CodArticol", "Price"])
        wb.close()


class TestResolveInputPath(unittest.TestCase):

    @patch.dict(os.environ, {"SCRAPER_PATH": "output/scraper_results.xlsx", "SCRAPER_INPUT": "input_articles.xlsx"})
    def test_resolves_correctly(self):
        result = resolve_input_path()
        self.assertEqual(result, os.path.join("output", "input_articles.xlsx"))

    @patch.dict(os.environ, {"SCRAPER_PATH": "results.xlsx", "SCRAPER_INPUT": "input.xlsx"})
    def test_no_directory_in_scraper_path(self):
        result = resolve_input_path()
        self.assertEqual(result, "input.xlsx")

    @patch.dict(os.environ, {"SCRAPER_PATH": "output/scraper_results.xlsx", "SCRAPER_INPUT": ""}, clear=False)
    def test_missing_scraper_input_raises(self):
        with self.assertRaises(ValueError):
            resolve_input_path()


class TestProcessScrapeIntegration(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    @patch("scrape.run_scrapers")
    @patch("scrape.excel.read_excel_input")
    @patch("scrape.resolve_input_path")
    @patch("scrape.excel.generate_timestamped_filename")
    @patch("scrape.load_scraper_config")
    def test_full_flow(self, mock_config, mock_ts_fn, mock_resolve, mock_read, mock_run):
        mock_config.return_value = {
            "input": {"id_column": "CodArticol"},
            "eoriginal": {"columns": {}},
        }
        mock_resolve.return_value = os.path.join(self.tmpdir, "input.xlsx")

        # Create input Excel
        input_path = mock_resolve.return_value
        _create_excel(input_path, ["CodArticol", "Desc"], [["A100", "Widget"], ["B200", "Gadget"]])
        mock_read.return_value = (
            ["A100", "B200"],
            [{"CodArticol": "A100", "Desc": "Widget"}, {"CodArticol": "B200", "Desc": "Gadget"}],
            ["CodArticol", "Desc"],
        )

        mock_run.return_value = [
            {"CodArticol": "A100", "PretNormal": 10.5, "PretStocAD": 12.0},
            {"CodArticol": "B200", "PretNormal": 20.0, "PretStocAD": None},
        ]

        out_path = os.path.join(self.tmpdir, "output.xlsx")
        mock_ts_fn.return_value = out_path

        from scrape import process_scrape
        success, created_path, count = process_scrape()

        self.assertTrue(success)
        self.assertEqual(created_path, out_path)
        self.assertEqual(count, 2)

        # Verify the output file
        wb = load_workbook(out_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, ["CodArticol", "Desc", "PretNormal", "PretStocAD"])
        wb.close()

    @patch("scrape.run_scrapers")
    @patch("scrape.excel.export_to_excel")
    @patch("scrape.load_scraper_config")
    def test_explicit_search_terms_uses_old_path(self, mock_config, mock_export, mock_run):
        mock_config.return_value = {"eoriginal": {"columns": {}}}
        mock_run.return_value = [{"CodArticol": "A100", "PretNormal": 10.0}]
        mock_export.return_value = "/tmp/out.xlsx"

        from scrape import process_scrape
        success, path, count = process_scrape(search_terms=["A100"])

        self.assertTrue(success)
        mock_export.assert_called_once()


if __name__ == "__main__":
    unittest.main()
