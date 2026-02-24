import os
from dotenv import load_dotenv
from openpyxl import Workbook
from datetime import datetime
import yaml

load_dotenv()

EXCEL_PATH = os.getenv("EXCEL_PATH")

def get_column_order(config_path="field_config.yaml"):
    """
    Get the preferred column order from the config file.
    System columns (order_number, email_date) come first, followed by pattern fields.
    """
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    
    # Define system columns that should appear first
    system_fields = ['order_number', 'email_date']
    ordered_columns = []
    
    # Add system columns first
    for field in system_fields:
        if field in config:
            excel_column = config[field].get('excel_column')
            if excel_column:
                ordered_columns.append(excel_column)
    
    # Add pattern-matching fields
    for field, props in config.items():
        if field not in system_fields and props.get('pattern') and props.get('excel_column'):
            ordered_columns.append(props['excel_column'])
    
    return ordered_columns

def generate_timestamped_filename(base_path=None):
    """
    Generate a timestamped Excel filename based on the current date and time.
    Format: YYYY-MM-DD_HH-MM-SS_original_filename.xlsx
    """
    if base_path is None:
        base_path = os.getenv("EXCEL_PATH")
    if not base_path:
        raise ValueError("No base path provided and EXCEL_PATH not set in .env file.")
    
    # Get current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    
    # Extract directory and filename components
    dir_name = os.path.dirname(base_path)
    base_name = os.path.basename(base_path)
    name, ext = os.path.splitext(base_name)
    
    # Create timestamped filename
    timestamped_filename = f"{timestamp}_{name}{ext}"
    
    # Create full path
    timestamped_path = os.path.join(dir_name, timestamped_filename)
    
    # Ensure directory exists
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name)
    
    return timestamped_path

def export_to_excel(data, excel_path=None):
    """
    Creates a new Excel file with a timestamped filename for each run.
    Each dict in data is written as a row to the Excel file.
    Each dict's keys are column names; values are written to the corresponding columns.
    Creates the file and columns if they do not exist.
    """
    if not data:
        print("No data to export. Skipping Excel file creation.")
        return None
    
    # Generate timestamped filename for this run
    if excel_path is None:
        excel_path = generate_timestamped_filename()
    
    print(f"Creating Excel file: {excel_path}")

    # Get preferred column order from configuration
    preferred_columns = get_column_order()
    
    # Determine all columns present in the data
    all_columns = set()
    for row in data:
        all_columns.update(row.keys())
    
    # Create final column list: preferred columns first, then any extra columns
    columns = []
    # Add preferred columns that exist in data
    for col in preferred_columns:
        if col in all_columns:
            columns.append(col)
    
    # Add any remaining columns that weren't in the preferred list
    for col in sorted(all_columns):  # Sort for consistent ordering
        if col not in columns:
            columns.append(col)

    # Always create a new workbook with timestamped filename
    wb = Workbook()
    ws = wb.active
    ws.append(columns)

    for row in data:
        # Ensure order matches columns
        ws.append([row.get(col, "") for col in columns])

    wb.save(excel_path)
    print(f"Excel file created successfully with {len(data)} rows and columns: {', '.join(columns)}")
    return excel_path


def read_excel_input(file_path, id_column):
    """
    Read an input Excel file and return search terms plus the original data.
    Returns (search_terms, rows, columns) where:
    - search_terms: list of non-empty ID strings
    - rows: list of dicts (one per row, only rows with non-empty ID)
    - columns: list of header names in original order
    """
    from openpyxl import load_workbook

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Input file not found: {file_path}")

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if headers is None:
        wb.close()
        raise ValueError(f"Input file is empty: {file_path}")

    columns = [str(h) for h in headers]

    if id_column not in columns:
        wb.close()
        raise ValueError(f"ID column '{id_column}' not found in input file. Available columns: {columns}")

    id_idx = columns.index(id_column)
    search_terms = []
    original_rows = []

    for row in rows_iter:
        id_val = row[id_idx]
        if id_val is None or str(id_val).strip() == "":
            continue
        row_dict = {columns[i]: row[i] for i in range(len(columns))}
        original_rows.append(row_dict)
        search_terms.append(str(id_val).strip())

    wb.close()
    return search_terms, original_rows, columns


def export_scraper_excel(original_rows, original_columns, scraper_results, id_column, excel_path):
    """
    Merge scraper results into original rows and export to Excel.
    - original_rows: list of dicts from input file
    - original_columns: list of header names in original order
    - scraper_results: list of dicts from scrapers (each has id_column key + price cols)
    - id_column: name of the ID column used for matching
    - excel_path: output file path
    """
    if not original_rows:
        print("No data to export. Skipping Excel file creation.")
        return None

    # Build lookup: id_value -> merged scraper dict
    scraper_lookup = {}
    for result in scraper_results:
        id_val = str(result.get(id_column, "")).strip()
        if id_val:
            if id_val in scraper_lookup:
                scraper_lookup[id_val].update(result)
            else:
                scraper_lookup[id_val] = dict(result)

    # Determine new columns from scraper results (not in original, preserving first-seen order)
    original_set = set(original_columns)
    new_columns = []
    seen = set()
    for result in scraper_results:
        for key in result:
            if key not in original_set and key not in seen:
                new_columns.append(key)
                seen.add(key)

    all_columns = original_columns + new_columns

    wb = Workbook()
    ws = wb.active
    ws.append(all_columns)

    for row in original_rows:
        id_val = str(row.get(id_column, "")).strip()
        scraper_data = scraper_lookup.get(id_val, {})
        combined = []
        for col in all_columns:
            if col in original_set:
                combined.append(row.get(col, ""))
            else:
                combined.append(scraper_data.get(col, ""))
        ws.append(combined)

    wb.save(excel_path)
    print(f"Scraper Excel created: {excel_path} ({len(original_rows)} rows, columns: {', '.join(all_columns)})")
    return excel_path
