import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

load_dotenv()

EXCEL_PATH = os.getenv("EXCEL_PATH")

def export_to_excel(data, excel_path=EXCEL_PATH):
    """
    Appends each dict in data to the Excel file at excel_path.
    Each dict's keys are column names; values are written to the corresponding columns.
    Creates the file and columns if they do not exist.
    """
    if not data:
        return
    if not excel_path:
        raise ValueError("EXCEL_PATH not set in .env file.")

    # Determine all columns from all dicts in data
    columns = set()
    for row in data:
        columns.update(row.keys())
    columns = list(columns)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        # Check if columns exist, else add header
        if ws.max_row == 0 or ws.max_column == 0:
            ws.append(columns)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(columns)

    for row in data:
        # Ensure order matches columns
        ws.append([row.get(col, "") for col in columns])

    wb.save(excel_path)
